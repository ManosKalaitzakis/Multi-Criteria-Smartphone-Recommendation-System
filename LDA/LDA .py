import re # We clean text using regex
import csv # To read the csv
from collections import defaultdict
import json
from gensim.models import CoherenceModel
import os
import gensim
from gensim.models.wrappers import LdaMallet
from pprint import pprint
import matplotlib.pyplot as plt
from gensim.models.wrappers import LdaMallet
import pyLDAvis.gensim
from collections import defaultdict # For accumlating values
from nltk.corpus import stopwords # To remove stopwords
from gensim import corpora # To create corpus and dictionary for the LDA model
from gensim.models import LdaModel # To use the LDA model
import pyLDAvis.gensim # To visualise LDA model effectively
import pandas as pd
import nltk
import unicodedata as ud
import numpy as np
os.environ.update({'MALLET_HOME':r'C:/mallet-2.0.8/'})
mallet_path = r'C:/mallet-2.0.8/bin/mallet'
def levenshtein_ratio_and_distance(s, t, ratio_calc = False):
    """ levenshtein_ratio_and_distance:
        Calculates levenshtein distance between two strings.
        If ratio_calc = True, the function computes the
        levenshtein distance ratio of similarity between two strings
        For all i and j, distance[i,j] will contain the Levenshtein
        distance between the first i characters of s and the
        first j characters of t
    """

    rows = len(s)+1
    cols = len(t)+1
    distance = np.zeros((rows,cols),dtype = int)

    for i in range(1, rows):
        for k in range(1,cols):
            distance[i][0] = i
            distance[0][k] = k

    for col in range(1, cols):
        for row in range(1, rows):
            if s[row-1] == t[col-1]:
                cost = 0
            else:
                if ratio_calc == True:
                    cost = 2
                else:
                    cost = 1
            distance[row][col] = min(distance[row-1][col] + 1,      # Cost of deletions
                                 distance[row][col-1] + 1,          # Cost of insertions
                                 distance[row-1][col-1] + cost)     # Cost of substitutions
    if ratio_calc == True:
        # Computation of the Levenshtein Distance Ratio
        Ratio = ((len(s)+len(t)) - distance[row][col]) / (len(s)+len(t))
        return Ratio
    else:
        # print(distance) # Uncomment if you want to see the matrix showing how the algorithm computes the cost of deletions,
        # insertions and/or substitutions
        # This is the minimum number of edits needed to convert string a to string b
        return "The strings are {} edits away".format(distance[row][col])
def remove_tonoi(s):
    d = {ord('\N{COMBINING ACUTE ACCENT}'): None}
    aa = ud.normalize('NFD', s).upper().translate(d)
    return aa

nltk.download('stopwords')
column1=[]
from openpyxl import load_workbook
reviews=[]
def excel_column_name(n):
    """Number to Excel-style column name, e.g., 1 = A, 26 = Z, 27 = AA, 703 = AAA."""
    name = ''
    while n > 0:
        n, r = divmod (n - 1, 26)
        name = chr(r + ord('A')) + name
    return name
wb = load_workbook("kinhta25.xlsx")  # Work Book
ws = wb.get_sheet_by_name('Sheet1')  # Work Sheet
for k in range(2,280*6,6):
    column = ws[excel_column_name(k)]  # Column
    reviews = reviews+ [column[x].value for x in range(0,len(column))]
    #column1.append(column_list)
reviews = [re.sub(r'[^\w\s]',' ',str(item)) for item in reviews]  #remove all punctuations
reviews = [remove_tonoi(item)for item in reviews]
for i in reviews:
    if (len(i))<40:
        reviews.remove(i)
k=0
for a in reviews:
    if a!='NONE':
        k+=1
        print (a)
print('keimena kritikwn:',k)
stopwords1 = stopwords.words('greek') #C:\Users\User\AppData\Roaming\nltk_data\corpora\stopwords

stopwords2 = set(line.strip() for line in open('stopwords_el.txt', encoding="utf8"))# apo arxeio
stopwords=stopwords2.union(stopwords1)

texts = [[word for word in document.lower().split() if word not in stopwords] for document in reviews]

frequency = defaultdict(int)
for text in texts:
    for token in text:
         frequency[token] += 1
aux = [(frequency[key], key) for key in frequency]
aux.sort()
aux.reverse()

texts = [[token for token in text if frequency[token] >10] for text in texts] #if frequency[token] >30] agnooume tis lekseis kritikwn me <30 fores emfanishs
lista=defaultdict(list)
banlist=[]
dictionary = corpora.Dictionary(texts)
Safelist=['χρωματα','χρηματα','iphone','phone']
for count1,w1 in dictionary.items():
        for count2,w2 in dictionary.items():
            Ratio = levenshtein_ratio_and_distance(w1,w2,ratio_calc = True)
            if Ratio>0.845 and Ratio!=1 and w1 not in banlist and w1 not in Safelist and w1  not in Safelist:

                lista[w1].append(w2)




                if w1 in lista:
                    banlist.append(w2)



for a in lista:
    print(a, '_')
    for i in lista[a]:
        print('\t',i)

print('\n')
print(banlist)
'''

'''
counter=0
counter1=-1
counter3=-1
for i in texts:
    counter3+=1
    for k in i:
        counter1+=1

        for a, y in lista.items():
            if k in y:


                texts[counter3][counter1]=a

    counter1=-1
counter3=-1

texts = [[token for token in text if frequency[token] >20] for text in texts]
#print(texts)
frequency1 = defaultdict(int)
for text in texts:
    for token in text:
         frequency1[token] += 1
aux = [(frequency1[key], key) for key in frequency]
aux.sort()
aux.reverse()
x=0
for leksi,suxnothta in aux:
    x+=1
    print(leksi,'___',suxnothta)
    if x==40: break

dictionary = corpora.Dictionary(texts)
'''for i,y in dictionary.items():

    #print(dictionary[i])
    #print(lista[i])
    #print(k[i])
    try:
        #print(lista[i])
        #print(i,'_____')
        print(y,'-')
        #print(i,'---')
        print(lista[y],"___")
        #print(dictionary[i],'___',i,'___',lista[i])
        if dictionary[i] in lista:
            y=lista[i]
            print(i,'))))()))(")"()()()(()')
    except:
        pass
        '''



corpus = [dictionary.doc2bow(text) for text in texts]#30 krit 1200 freq kai 3 8emata
print(type(texts))
with open('texts.txt', 'w') as f:
    f.write(json.dumps(texts))

dictionary1= dict(list(dictionary.items())[1:100]) # na allaksw to 1o dictionary1 se dictionary
for a in dictionary1:
    print (a)


NUM_TOPICS =10

ldamodel = LdaModel(corpus, num_topics = NUM_TOPICS, id2word=dictionary, passes=10000, alpha=0.7,eta=0.1, per_word_topics= True) #This might take some time.
topics = ldamodel.show_topics()
modelname=str(NUM_TOPICS)+'_topics_model'+'.model'


for topic in topics:
    print(topic)
word_dict = {};

for i in range(NUM_TOPICS):
    words = ldamodel.show_topic(i, topn = 15)
    word_dict['Topic # ' + '{:02d}'.format(i+1)] = [i[0] for i in words]

pd.DataFrame(word_dict)
#ks.show
#pyLDAvis.save_html(ldamodel, 'lda.html')

#lda_display_Mallet= pyLDAvis.gensim.prepare(ldamallet,corpus,dictionary,sort_topics=False)
#pyLDAvis.show(lda_display_Mallet)
lda_display = pyLDAvis.gensim.prepare(ldamodel, corpus, dictionary, sort_topics=False)
pyLDAvis.show(lda_display)
#pyLDAvis.save_html(ldamodel, 'lda.html')